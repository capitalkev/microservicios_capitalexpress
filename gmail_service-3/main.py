import os
import base64
import mimetypes
import io
import pandas as pd
from fastapi import FastAPI, Request, HTTPException
from email.message import EmailMessage
from dotenv import load_dotenv
from collections import defaultdict
from pydantic import BaseModel
from typing import List, Optional

# --- Importaciones de Google ---
from google.cloud import storage
from google.oauth2.credentials import Credentials
from google.oauth2.service_account import Credentials as ServiceAccountCredentials
from googleapiclient.discovery import build

# --- Importaciones para crear Excel ---
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

load_dotenv()

# --- Configuración ---
USER_TOKEN_FILE = 'token.json'
SERVICE_ACCOUNT_FILE = os.getenv("GOOGLE_APPLICATION_CREDENTIALS")
SENDER_USER_ID = 'kevin.gianecchine@capitalexpress.cl'
SCOPES = ['https://www.googleapis.com/auth/gmail.send', 'https://www.googleapis.com/auth/devstorage.read_only']

FIXED_CC_LIST = [
    'kevin.gianecchine@capitalexpress.cl', 'jenssy.huaman@capitalexpress.pe',
    'jakeline.quispe@capitalexpress.pe', 'jhonny.celay@capitalexpress.pe',
    'kevin.tupac@capitalexpress.cl'
]

RUC_GLORIA = [
    "20100190797", "20600679164", "20312372895", "20524088739", "20467539842",
    "20506475288", "20418453177", "20512613218", "20115039262", "20100814162",
    "20518410858", "20101927904", "20479079006", "20100223555", "20532559147",
    "20487268870", "20562613545", "20478963719", "20481694907", "20454629516",
    "20512415840", "20602903193", "20392965191", "20601225639", "20547999691",
    "20600180631", "20116225779", "20131823020", "20601226015", "20131867744",
    "20603778180", "20131835621", "20511866210", "20481640483"
]

app = FastAPI(title="Servicio de Gmail Híbrido Avanzado")

class InvoiceData(BaseModel):
    document_id: Optional[str] = None
    issue_date: Optional[str] = None
    due_date: Optional[str] = None
    currency: Optional[str] = None
    total_amount: Optional[float] = 0.0
    net_amount: Optional[float] = 0.0
    debtor_name: Optional[str] = None
    debtor_ruc: Optional[str] = None
    client_name: Optional[str] = None
    client_ruc: Optional[str] = None

# --- Funciones de Autenticación ---
def get_user_credentials():
    creds = None
    if os.path.exists(USER_TOKEN_FILE):
        creds = Credentials.from_authorized_user_file(USER_TOKEN_FILE, SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            from google.auth.transport.requests import Request as GoogleRequest
            creds.refresh(GoogleRequest())
        else:
            raise Exception("No se encontraron credenciales de usuario válidas.")
    return creds

def get_storage_client():
    if not SERVICE_ACCOUNT_FILE or not os.path.exists(SERVICE_ACCOUNT_FILE):
        raise Exception("No se encontró el archivo de cuenta de servicio.")
    sa_creds = ServiceAccountCredentials.from_service_account_file(SERVICE_ACCOUNT_FILE)
    return storage.Client(credentials=sa_creds)

# --- Función para crear el Excel de Gloria ---
def create_gloria_excel(invoice_data_list: List[InvoiceData]) -> (str, bytes):
    if not invoice_data_list:
        return None, None
    first_invoice = invoice_data_list[0]
    proveedores_juntos = '\n'.join(sorted(list(set(inv.debtor_name for inv in invoice_data_list if inv.debtor_name))))
    data = {
        'FACTOR': [first_invoice.client_name],
        'FECHA DE ENVIO': [pd.to_datetime('today').strftime('%d/%m/%Y')],
        'RUC PROVEEDOR': ['\n'.join(inv.debtor_ruc for inv in invoice_data_list)],
        'PROVEEDOR': [proveedores_juntos],
        'RUC CLIENTE': [first_invoice.client_ruc],
        'CLIENTE': [first_invoice.client_name],
        'FECHA DE EMISION': [pd.to_datetime(first_invoice.issue_date, errors='coerce').strftime('%d/%m/%Y')],
        'NUM FACTURA': ['\n'.join(inv.document_id for inv in invoice_data_list)],
        'IMPORTE NETO PAGAR': [sum(inv.net_amount for inv in invoice_data_list)],
        'MONEDA': [first_invoice.currency],
        'FECHA DE VENCIMIENTO': [pd.to_datetime(first_invoice.due_date, errors='coerce').strftime('%d/%m/%Y')]
    }
    df = pd.DataFrame(data)
    output_buffer = io.BytesIO()
    with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Facturas', index=False)
        worksheet = writer.sheets['Facturas']
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        thin_border_side = Side(border_style="thin", color="000000")
        cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        center_alignment = Alignment(horizontal='center', vertical='center')
        right_alignment = Alignment(horizontal='right', vertical='center')
        wrap_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = cell_border
                cell.alignment = center_alignment
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
        for col_idx, column_cells in enumerate(worksheet.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for cell in column_cells:
                if column_letter in ['C', 'D', 'H'] and not cell.row == 1:
                    cell.alignment = wrap_alignment
                if '\n' in str(cell.value):
                    max_length = max(len(line) for line in str(cell.value).split('\n'))
                elif len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        for cell in worksheet['I'][1:]:
            cell.number_format = '#,##0.00'
            cell.alignment = right_alignment
    excel_bytes = output_buffer.getvalue()
    filename = f"Facturas_{first_invoice.client_name.replace(' ', '_')}.xlsx"
    return filename, excel_bytes

# --- Función para Crear el HTML ---
def create_html_body(invoice_data_list: List[InvoiceData]) -> str:
    first_invoice = invoice_data_list[0]
    client_name = first_invoice.client_name
    client_ruc = first_invoice.client_ruc
    data_for_df = [invoice.dict() for invoice in invoice_data_list]
    df = pd.DataFrame(data_for_df)
    df['total_amount'] = df.apply(lambda row: f"{row.get('currency', '')} {float(row.get('total_amount', 0)):,.2f}".strip(), axis=1)
    df['net_amount'] = df.apply(lambda row: f"{row.get('currency', '')} {float(row.get('net_amount', 0)):,.2f}".strip(), axis=1)
    df['due_date'] = pd.to_datetime(df['due_date'], errors='coerce').dt.strftime('%d/%m/%Y')
    df_display = df.rename(columns={
        'debtor_ruc': 'RUC Deudor', 'debtor_name': 'Nombre Deudor', 'document_id': 'Documento',
        'total_amount': 'Monto Factura', 'net_amount': 'Monto Neto', 'due_date': 'Fecha de Pago'
    })
    display_columns = ['RUC Deudor', 'Nombre Deudor', 'Documento', 'Monto Factura', 'Monto Neto', 'Fecha de Pago']
    tabla_html = df_display[display_columns].to_html(index=False, border=1, justify='left', classes='invoice_table')
    mensaje_html = f"""
    <!DOCTYPE html>
    <html lang="es">
    <head>
    <meta charset="UTF-8">
    <style>
        body {{ font-family: Arial, Helvetica, sans-serif; font-size: 13px; color: #000; }}
        .container {{ max-width: 800px; }} p, li {{ line-height: 1.5; }} ol {{ padding-left: 30px; }}
        table.invoice_table {{ width: 100%; border-collapse: collapse; margin-top: 15px; margin-bottom: 20px; }}
        table.invoice_table th, table.invoice_table td {{ border: 1px solid #777; padding: 6px; text-align: left; font-size: 12px; }}
        table.invoice_table th {{ background-color: #f0f0f0; font-weight: bold; }}
        .disclaimer {{ font-style: italic; font-size: 11px; margin-top: 25px; }}
    </style>
    </head>
    <body>
    <div class="container">
        <p>Estimados señores,</p>
        <p>Por medio de la presente, les informamos que los señores de <strong>{client_name}</strong>, nos han transferido la(s) siguiente(s) factura(s) negociable(s). Solicitamos su amable confirmación sobre los siguientes puntos:</p>
        <ol>
            <li>¿La(s) factura(s) ha(n) sido recepcionada(s) conforme con sus productos o servicios?</li>
            <li>¿Cuál es la fecha programada para el pago de la(s) misma(s)?</li>
            <li>Por favor, confirmar el Monto Neto a pagar, considerando detracciones, retenciones u otros descuentos.</li>
        </ol>
        <p><strong>Detalle de las facturas:</strong></p>
        <p>Cliente: {client_name}<br>
        RUC Cliente: {client_ruc}
        </p>
        {tabla_html}
        <p>Agradecemos de antemano su pronta respuesta. Con su confirmación, procederemos a la anotación en cuenta en CAVALI.</p>
        <p class="disclaimer">"Sin perjuicio de lo anteriormente mencionado, nos permitimos recordarles que toda acción tendiente a 
        simular la emisión de la referida factura negociable o letra para obtener un beneficio a título personal o a favor de la otra 
        parte de la relación comercial, teniendo pleno conocimiento de que la misma no proviene de una relación comercial verdadera, 
        se encuentra sancionada penalmente como delito de estafa en nuestro ordenamiento jurídico.
        Asimismo, en caso de que vuestra representada cometa un delito de forma conjunta y/o en contubernio con el emitente de la factura, 
        dicha acción podría tipificarse como delito de asociación ilícita para delinquir, según el artículo 317 del Código Penal, por lo 
        que nos reservamos el derecho de iniciar las acciones penales correspondientes en caso resulte necesario"</p>
    </div>
    </body>
    </html>
    """
    return mensaje_html

# --- Endpoint Principal ---
@app.post("/gmail")
async def send_verification_email(request: Request):
    try:
        user_creds = get_user_credentials()
        storage_client = get_storage_client()
        gmail_service = build('gmail', 'v1', credentials=user_creds)

        data = await request.json()
        pdf_paths = data.get("pdf_paths", [])
        parser_results = data.get("parsed_invoice_data", {}).get("results", [])
        emails_from_excel = data.get("recipient_emails")
        user_email = data.get("user_email")

        if not parser_results or not emails_from_excel:
            raise HTTPException(status_code=400, detail="Faltan datos de facturas o correos de destinatarios.")
        
        cc_list = set(FIXED_CC_LIST)
        if user_email: cc_list.add(user_email)
        cc_string = ",".join(sorted(list(cc_list)))

        invoice_models = [InvoiceData(**res['parsed_invoice_data']) for res in parser_results if res.get('status') == 'SUCCESS']
        if not invoice_models:
            return {"status": "SKIPPED", "message": "No hay facturas válidas para enviar."}

        facturas_por_deudor = defaultdict(list)
        for invoice in invoice_models:
            facturas_por_deudor[invoice.debtor_ruc].append(invoice)

        # El bucle itera sobre cada RUC de deudor
        for ruc_deudor, facturas_grupo in facturas_por_deudor.items():
            message = EmailMessage()
            client_name = facturas_grupo[0].client_name

            # 1. Crear el cuerpo HTML (siempre es el mismo)
            html_body = create_html_body(facturas_grupo)
            message.add_alternative(html_body, subtype='html')
            
            print(f"DEBUG: Verificando RUC del DEUDOR. RUC: '{ruc_deudor}', ¿Está en la lista de Gloria?: {ruc_deudor in RUC_GLORIA}")

            if ruc_deudor in RUC_GLORIA:
                excel_filename, excel_bytes = create_gloria_excel(facturas_grupo)
                
                # PASO 2: Verificar si el archivo se generó en memoria
                if excel_bytes:
                    print(f"DEBUG: Archivo Excel CREADO para DEUDOR '{ruc_deudor}'. Tamaño: {len(excel_bytes)} bytes. Adjuntando...")
                    message.add_attachment(excel_bytes,
                                           maintype='application',
                                           subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                           filename=excel_filename)
                else:
                    print(f"ERROR: La función create_gloria_excel no devolvió datos para adjuntar.")
            
            # 3. Asignar destinatarios y asunto
            message['To'] = emails_from_excel
            message['Cc'] = cc_string
            message['Subject'] = f"Confirmación de Facturas Negociables - {client_name}"

            for pdf_path in pdf_paths:
                try:
                    bucket_name, blob_name = pdf_path.replace("gs://", "").split("/", 1)
                    blob = storage_client.bucket(bucket_name).blob(blob_name)
                    pdf_bytes = blob.download_as_bytes()
                    maintype, subtype = (mimetypes.guess_type(os.path.basename(pdf_path))[0] or "application/octet-stream").split('/')
                    message.add_attachment(pdf_bytes, maintype=maintype, subtype=subtype, filename=os.path.basename(pdf_path))
                except Exception as e:
                    print(f"ADVERTENCIA al adjuntar {pdf_path}: {e}")

            # 5. Enviar el correo
            encoded_message = base64.urlsafe_b64encode(message.as_bytes()).decode()
            gmail_service.users().messages().send(
                userId=SENDER_USER_ID, body={'raw': encoded_message}
            ).execute()
            print(f"Correo para deudor {ruc_deudor} enviado a: {emails_from_excel} con CC a: {cc_string}")

        return {"status": "SUCCESS", "message": "Correos de notificación enviados."}

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error interno en el servicio de Gmail: {str(e)}")